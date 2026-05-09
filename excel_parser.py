import datetime
import io

import openpyxl


def parse_invoice_excel(excel_bytes: bytes) -> list[dict]:
    """
    Lee un Excel con columnas CUIT, Importe, Fecha.
    Retorna lista de dicts listos para enviar a /api/facturar.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    # Detectar columnas por encabezado (fila 1)
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip().lower() if cell.value is not None else ""
        headers.append(val)

    col = {}
    for i, h in enumerate(headers):
        if "cuit" in h:
            col["cuit"] = i
        elif "importe" in h or "monto" in h or "total" in h:
            col["importe"] = i
        elif "fecha" in h:
            col["fecha"] = i

    missing = [k for k in ("cuit", "importe", "fecha") if k not in col]
    if missing:
        raise ValueError(f"Faltan columnas en el Excel: {', '.join(missing).upper()}")

    results = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas vacías
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        cuit_raw = row[col["cuit"]]
        importe_raw = row[col["importe"]]
        fecha_raw = row[col["fecha"]]

        # CUIT
        cuit = str(int(cuit_raw)) if isinstance(cuit_raw, float) else str(cuit_raw or "")
        cuit = cuit.replace("-", "").replace(" ", "").strip()

        # Importe
        try:
            importe = round(float(str(importe_raw).replace(",", ".")), 2)
        except (ValueError, TypeError):
            raise ValueError(f"Fila {row_num}: importe inválido '{importe_raw}'")

        # Fecha — puede venir como datetime (Excel la parsea sola) o string
        if isinstance(fecha_raw, (datetime.datetime, datetime.date)):
            fecha = fecha_raw.strftime("%d/%m/%Y")
        else:
            fecha = str(fecha_raw).strip() if fecha_raw else ""

        if not fecha:
            raise ValueError(f"Fila {row_num}: fecha vacía")

        results.append({
            "date":             fecha,
            "description":      "",
            "amount":           importe,
            "receptor_cuit":    cuit,
            "concepto":         2,
            "use_month_period": True,
        })

    return results
